#!/usr/bin/env python3
"""
Regenerate api/manual/roster-data.js from a fresh export of
"Real Estate Agent Management Dashboard.xlsx".

Usage:
    pip install openpyxl
    python3 scripts/build-manual-roster.py "path/to/Real Estate Agent Management Dashboard.xlsx"

This overwrites api/manual/roster-data.js (relative to the repo root, i.e.
one level up from this scripts/ folder) with the latest snapshot. Commit
the result the same way you'd commit any other code change — it's a plain
JS file with the data baked in, not a live spreadsheet connection.

Run this whenever the roster spreadsheet changes materially (new starters,
license renewals recorded, training hours updated, supplier contracts
renewed, etc.) — a monthly refresh alongside the existing "check License
Expiry column" routine (manual §2.1) works well.
"""
import sys
import json
import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency — run: pip install openpyxl")


def d(v):
    """Excel date -> ISO date string, or None for blank/zero cells."""
    if v is None or v == "" or v == 0:
        return None
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    return v


def build(xlsx_path: Path, out_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    ws = wb["Real Estate Agent Management Da"]
    roster = []
    for row in range(2, ws.max_row + 1):
        vals = {ws.cell(row=1, column=c).value: ws.cell(row=row, column=c).value
                for c in range(1, 22)}
        if not vals.get("Agent First Name"):
            continue
        licence = vals.get("Licence Number")
        roster.append({
            "firstName": vals.get("Agent First Name"),
            "surname": vals.get("Surname"),
            "email": vals.get("Email Address"),
            "mobile": vals.get("Mobile Number"),
            "dob": d(vals.get("Date of Birth")),
            "workAnniversary": d(vals.get("Work Anniversary")),
            "yearsWithColliers": vals.get("Years with Colliers"),
            "yearStartedRE": d(vals.get("Year Started in RE")),
            "experience": vals.get("Experience"),
            "supervisionLevel": vals.get("Supervision Level"),
            "supervisionFrequency": vals.get("Frequency of Supervision Meeting"),
            "jobTitle": vals.get("Job Title"),
            "licenceNumber": str(licence).strip() if licence not in (None, "") else None,
            "supervisionPlanStart": d(vals.get("Supervision_Plan_Start")),
            "reviewDate": d(vals.get("Review_Date")),
            "verifiableHours": vals.get("Verifiable Training Hours (Y/N)"),
            "nonVerifiableHours": vals.get("Non-Verifiable Training Hours"),
            "totalCpdHours": vals.get("Total CPD Hours"),
            "licenseExpiry": d(vals.get("License Expiry")),
            "activeListings": vals.get("Active Listings"),
            "address": vals.get("Address") or None,
        })

    ws2 = wb["Broker Contract Audits"]
    audits = []
    for row in range(2, ws2.max_row + 1):
        vals = [ws2.cell(row=row, column=c).value for c in range(1, 6)]
        if not vals[0]:
            continue
        audits.append({
            "firstName": vals[0], "surname": vals[1],
            "issuesIdentified": vals[2], "trainingGiven": vals[3], "comments": vals[4],
        })

    ws3 = wb["Project - Open Issues Register"]
    headers3 = [ws3.cell(row=1, column=c).value for c in range(1, 11)]
    issues = []
    for row in range(2, ws3.max_row + 1):
        vals = [ws3.cell(row=row, column=c).value for c in range(1, 11)]
        if not any(vals):
            continue
        issues.append(dict(zip(headers3, vals)))

    ws4 = wb["Suppliers & Sponsors"]
    headers4 = [ws4.cell(row=1, column=c).value for c in range(1, 14)]
    suppliers = []
    for row in range(2, ws4.max_row + 1):
        vals = [ws4.cell(row=row, column=c).value for c in range(1, 14)]
        if not any(vals):
            continue
        suppliers.append(dict(zip(headers4, vals)))

    # Supervision process text and REINZ awards are narrative, not tabular —
    # they rarely change, so they're kept here rather than re-scraped from
    # the "Supervision" / "REINZ Awards" sheets. Edit them directly below
    # (or in the generated file) if the wording changes.
    supervision_process = {
        "principles": [
            "Courtney & Nick have their Agent licence and can supervise brokers and complete the logs.",
            "Sam & Brynn do not have their branch managers but can be the line managers for their brokers.",
            "Jason (COO) meets with the team leader and the broker and completes the logs for the brokers, "
            "with the team leader facilitating the session — the team leader understands the broker's "
            "skill level and can give accurate feedback on what they need to improve on.",
            "All Word documents are kept in the broker's supervision folder and updated at the frequency "
            "required by the supervision matrix.",
        ],
        "tasks": [
            "Meet with team leaders and explain the process for supervision. Find out what level of "
            "experience each broker has and when their last 1:1 supervision meeting was.",
            "At a Thursday meeting, explain to the team the purpose of supervision and why records need "
            "updating — create a slide for this discussion point.",
            "Schedule the first supervision meeting with line managers, broker and JW.",
            "Complete each broker's supervision plan with their capabilities and issues, and update the "
            "plan's communication log.",
            "Ask the broker to sign the supervision plan, scan to PDF and file in their personal folder "
            "under Supervision.",
        ],
    }
    reinz_awards = {
        "note": "Maree is collating the information for these awards. Anna is registering our Sam and our "
                "office for these two awards. Mark feels we won't have big enough deals to win the Deal of "
                "the Year awards.",
        "categories": [
            "Individual Commercial & Industrial Salesperson of the Year",
            "Commercial & Industrial Office of the Year – Medium",
            "New Sales Deal of the Year — recognises a single, completed commercial or industrial "
            "property transaction. Eligible transactions are assessed on settlement date and verified by "
            "an independent auditor.",
            "New Commercial Leasing Deal of the Year — recognises a single completed commercial "
            "property leasing transaction that demonstrates scale and commercial value. Qualifying lease "
            "deals are verified by an independent auditor.",
        ],
    }

    snapshot_date = datetime.date.today().isoformat()
    out = []
    out.append("// AUTO-GENERATED from Real Estate Agent Management Dashboard.xlsx")
    out.append("// Regenerate with: python3 scripts/build-manual-roster.py <path-to-xlsx>")
    out.append(f"// Snapshot date: {snapshot_date}\n")
    out.append(f'export const DASHBOARD_SNAPSHOT_DATE = "{snapshot_date}";\n')
    out.append("export const ROSTER = " + json.dumps(roster, indent=2) + ";\n")
    out.append("export const BROKER_CONTRACT_AUDITS = " + json.dumps(audits, indent=2) + ";\n")
    out.append("export const OPEN_ISSUES_REGISTER = " + json.dumps(issues, indent=2, default=str) + ";\n")
    out.append("export const SUPPLIERS_SPONSORS = " + json.dumps(suppliers, indent=2, default=str) + ";\n")
    out.append("export const REINZ_AWARDS = " + json.dumps(reinz_awards, indent=2) + ";\n")
    out.append("export const SUPERVISION_PROCESS = " + json.dumps(supervision_process, indent=2) + ";\n")

    out_path.write_text("\n".join(out), encoding="utf-8")
    print(f"Wrote {out_path} — {len(roster)} roster rows, {len(audits)} audits, "
          f"{len(issues)} open issues, {len(suppliers)} suppliers.")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit(f"Usage: python3 {sys.argv[0]} <path-to-xlsx>")
    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        sys.exit(f"File not found: {xlsx_path}")
    repo_root = Path(__file__).resolve().parent.parent
    out_path = repo_root / "api" / "manual" / "roster-data.js"
    build(xlsx_path, out_path)
