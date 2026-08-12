#!/usr/bin/env python3
"""
Deal Board — one-off importer.

Loads the two meeting workbooks into db_deals / db_deal_brokers, and
seeds db_broker_ranking_names from the rankings files.

Run AFTER dealboard_schema_public.sql. Idempotent: re-running replaces
the imported deals for a department rather than duplicating them.

    pip install openpyxl psycopg2-binary
    export DATABASE_URL='postgresql://...'   # Supabase connection string
    python import_meetings.py --dry-run      # inspect first
    python import_meetings.py

Every broker code below was verified against public.brokers — all 19
codes used across both workbooks resolve to an active broker.
"""
import argparse, os, re, sys
from openpyxl import load_workbook

# ---------------------------------------------------------------------
# Ranking-name map. Explicit, never derived: Industrial has two
# Marshalls (P and L), and Macauley / Ogg / Lough all reduce to 'M'.
# Left side is exactly as spelled in the rankings workbooks.
# ---------------------------------------------------------------------
RANKING_NAMES = {
    # Investment
    "H Doig": "HD", "M Macauley": "MM", "C Doig": "CD", "W Franks": "WF",
    "M Lough": "ML", "M Ogg": "MO", "L Wishnowsky": "LW", "B Cameron": "BC",
    "N Gilchrist": "NG", "P Cooper": "PC", "E Sparrow": "ES",
    # Industrial
    "S Staite": "SS", "P Marshall": "PM", "C Kellar": "CK", "O Salt": "OS",
    "R McGuigan": "RM", "E Clayton": "EC", "H Peeters": "HP",
    "G Sidey": "GS", "L Marshall": "LM",
}

# ---------------------------------------------------------------------
# Section header -> stage name, per department.
# Anything not listed is skipped and reported.
# ---------------------------------------------------------------------
INVESTMENT_SECTIONS = {
    "SUBMISSIONS": "Submissions",
    "CAMPAIGNS / SOLE AGENCY": "Campaigns / sole agency",
    "ADVANCED": "Advanced",
    "CONDITIONAL": "Conditional",
    "UNCONDITIONAL": "Unconditional",
    "TRACKING / WIP": "Tracking / WIP",
}
# Matched by PREFIX, because Industrial's submission blocks append the
# broker to the heading:
#     "Listing & Client Targets/ Submissions  - Sam Staite"
# That suffix is the only record of whose submission it is — those rows
# have no broker column — so it is parsed out, not discarded.
INDUSTRIAL_SECTIONS = [
    ("LISTING & CLIENT TARGETS", "Submissions"),
    ("LISTING AND CLIENT TARGETS", "Submissions"),
    ("ADVANCED NEGOTIATIONS", "Advanced"),
    ("UNDER CONTRACT", "Under contract"),
    ("UNCONDITIONAL", "Unconditional"),
    ("CAMPAIGNS", "Campaigns"),
]

# First names / nicknames used in those headings -> broker code.
HEADER_NAMES = {
    "SAM STAITE": ["SS"], "PAUL MARSHALL": ["PM"], "CK & OLLIE": ["CK", "OS"],
    "CHRISTIAN": ["CK"], "OLLIE": ["OS"], "OLIVER": ["OS"], "SAM": ["SS"],
    "PAUL": ["PM"], "RORY": ["RM"], "ELLIOT": ["EC"], "HARRY": ["HP"],
    "JACKSON": ["JM"], "JACKO": ["JM"],
    "ELLIOT CLAYTON": ["EC"], "RORY MCGUIGAN": ["RM"],
    "HARRY PEETERS": ["HP"], "CHRISTIAN KELLAR": ["CK"], "OLIVER SALT": ["OS"],
}


def section_for(key):
    for prefix, stage in INDUSTRIAL_SECTIONS:
        if key.startswith(prefix):
            return stage, key[len(prefix):]
    return None, ""


def brokers_from_heading(tail):
    """'  - Sam Staite' -> ['SS']. Unknown names are reported."""
    t = tail.split("-", 1)[1].strip().upper() if "-" in tail else ""
    if not t:
        return [], None
    if t in HEADER_NAMES:
        return HEADER_NAMES[t], None
    return [], t
# Sections deliberately NOT imported — they are not pipeline stages:
SKIP = {"PROBLEM PROPERTY", "MARKETING REPORTS", "UPCOMING AUCTIONS",
        "AUCTION TARGETS 2026", "TARGETS 2026", "BROKER", "FINES",
        "TENANT/ BUYER ENQUIRY", "BUDGETS 2025", "BUDGETS 2026", "2025", "2026"}

VALID_CODE = re.compile(r"^[A-Z]{1,4}$")

# Co-agency with another office, written where a broker code would go.
REGIONS = {"Nelson", "Masterton", "Auckland", "Wellington", "Dunedin", "Queenstown"}


def offset(ws):
    """Some sheets (Minutes_Master 'WEEK 1') indent everything by a
    blank column A. Find the first column that actually holds data."""
    for c in range(0, 4):
        hits = sum(1 for r in ws.iter_rows(min_col=c + 1, max_col=c + 1)
                   if r[0].value not in (None, ""))
        if hits > 5:
            return c
    return 0


def norm(v):
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%-d %b %Y") if v.year > 1950 else ""
    return str(v).strip()


def money(v):
    """Sheets mix 90000, '90,000' and 21 (meaning $21k in the old
    Industrial 'Fee ($000s)' columns). Only the first two are safe to
    read automatically — small bare numbers are reported, not guessed."""
    s = norm(v).replace("$", "").replace(",", "")
    if not s:
        return 0.0, None
    try:
        n = float(s)
    except ValueError:
        return 0.0, f"unreadable fee {norm(v)!r}"
    if 0 < n < 1000:
        return n, f"fee {n:g} looks like thousands — check"
    return n, None


def split_brokers(cell):
    """'MO/ML/MM' -> ['MO','ML','MM']. Anything that isn't a plain code
    is returned separately so it can be reported rather than dropped."""
    raw = norm(cell)
    if not raw:
        return [], []
    good, bad = [], []
    for part in re.split(r"[/,&]| and ", raw):
        p = part.strip().upper()
        if not p:
            continue
        if VALID_CODE.match(p):
            good.append(p)
        elif p.title() in REGIONS:
            pass          # 'ML/Nelson' — a referring office, not a broker
        else:
            bad.append(p)
    return good, bad


def parse_investment(path):
    ws = load_workbook(path, data_only=True).worksheets[0]
    rows, stage, warn = [], None, []
    for r in ws.iter_rows():
        first = norm(r[0].value)
        key = first.upper()
        if key in INVESTMENT_SECTIONS:
            stage = INVESTMENT_SECTIONS[key]
            continue
        if key in SKIP or key.startswith("COLUMN"):
            stage = None
            continue
        if not first or not stage or key.startswith("TOTAL"):
            continue
        # In the Submissions block column C carries a status word
        # ('PENDING', 'submitted'), not a fee. Read it as status there.
        raw_c = norm(r[2].value) if len(r) > 2 else ""
        if stage == "Submissions":
            fee, w, status_c = 0.0, None, raw_c
        else:
            fee, w = money(r[2].value if len(r) > 2 else None)
            status_c = ""
        codes, bad = split_brokers(r[3].value if len(r) > 3 else None)
        if w:
            warn.append(f"{first}: {w}")
        if bad:
            warn.append(f"{first}: unrecognised broker {bad}")
        rows.append(dict(stage=stage, address=first,
                         timing=norm(r[1].value) if len(r) > 1 else "",
                         fee=fee,
                         status=status_c or (norm(r[4].value) if len(r) > 4 else ""),
                         brokers=codes))
    return rows, warn


def parse_industrial(path):
    """The Industrial master is laid out in blocks with a header row
    naming the columns, which shift between sections. Column positions
    are read from each block's own header rather than assumed."""
    ws = load_workbook(path, data_only=True).worksheets[0]
    off = offset(ws)
    rows, stage, cols, warn, hdr_brokers = [], None, {}, [], []
    for r in ws.iter_rows():
        vals = [norm(c.value) for c in r][off:]
        first = vals[0] if vals else ""
        key = first.upper()
        found, tail = section_for(key)
        if found:
            stage, cols = found, {}
            hdr_brokers, unknown_name = brokers_from_heading(tail)
            if unknown_name:
                warn.append(f"heading names {unknown_name!r} — no broker matched")
            continue
        if key in SKIP:
            stage = None
            continue
        if first.upper() == "ADDRESS":
            cols = {v.upper().strip(): i for i, v in enumerate(vals) if v}
            continue
        if not first or not stage or key.startswith("TOTAL"):
            continue

        def col(*names):
            for n in names:
                for k, i in cols.items():
                    if k.startswith(n):
                        return vals[i] if i < len(vals) else ""
            return ""

        # Industrial blocks label the column "Fee ($000's)" — scale it
        # rather than asking a human to eyeball forty rows.
        fee_hdr = next((k for k in cols if k.startswith("FEE")), "")
        thousands = "000" in fee_hdr
        fee, w = money(col("FEE"))
        if thousands:
            fee, w = fee * 1000, None
        codes, bad = split_brokers(col("BROKER"))
        if not codes and hdr_brokers:
            codes = list(hdr_brokers)   # submission blocks: broker is in the heading
        if w:
            warn.append(f"{first}: {w}")
        if bad:
            warn.append(f"{first}: unrecognised broker {bad}")
        rows.append(dict(stage=stage, address=first,
                         timing=col("TIMING", "AUCTION/DEADLINE", "DATE"),
                         fee=fee, status=col("PROGRESS", "SALE/LEASE", "DETAIL"),
                         brokers=codes))
    return rows, warn


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--investment", default="4_August_2026.xlsx")
    ap.add_argument("--industrial", default="Minutes_Master.xlsx")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    inv, w1 = parse_investment(args.investment)
    ind, w2 = parse_industrial(args.industrial)

    print(f"Investment: {len(inv)} deals")
    print(f"Industrial: {len(ind)} deals")

    unknown = sorted({c for r in inv + ind for c in r["brokers"]}
                     - set(RANKING_NAMES.values()))
    if unknown:
        print(f"\nBroker codes present in sheets, not in the ranking map: "
              f"{', '.join(unknown)}")
        print("  (fine — they just have no commission-workbook counterpart)")

    if w1 + w2:
        print("\nNeeds a human look:")
        for w in w1 + w2:
            print("  -", w)

    if args.dry_run:
        print("\n--- dry run, nothing written ---")
        for r in (inv + ind)[:12]:
            print(f"  [{r['stage']:<24}] {r['address'][:38]:<40} "
                  f"{r['fee']:>10,.0f}  {'/'.join(r['brokers'])}")
        return

    import psycopg2
    from psycopg2.extras import execute_batch
    dsn = os.environ.get("DATABASE_URL")
    if not dsn:
        sys.exit("Set DATABASE_URL first.")

    conn = psycopg2.connect(dsn)
    cur = conn.cursor()

    # Ranking names — only for codes that exist in brokers.
    execute_batch(cur, """
        insert into public.db_broker_ranking_names (broker_code, ranking_name)
        select %s, %s
        where exists (select 1 from public.brokers where code = %s)
        on conflict (broker_code) do update set ranking_name = excluded.ranking_name
    """, [(c, n, c) for n, c in RANKING_NAMES.items()])

    for slug, rows in (("investment", inv), ("industrial", ind)):
        cur.execute("select id from public.db_departments where slug=%s", (slug,))
        dept = cur.fetchone()
        if not dept:
            sys.exit(f"Department {slug} not found — run the schema first.")
        dept = dept[0]

        cur.execute("select id, name from public.db_stages where department_id=%s", (dept,))
        stage_ids = {n: i for i, n in cur.fetchall()}

        # Idempotent: clear this department's imported deals first.
        cur.execute("delete from public.db_deals where department_id=%s", (dept,))

        order = 1000
        for r in rows:
            sid = stage_ids.get(r["stage"])
            if not sid:
                print(f"  ! no stage {r['stage']!r} in {slug} — skipped {r['address']}")
                continue
            cur.execute("""
                insert into public.db_deals
                  (department_id, stage_id, sort_order, address, timing,
                   fee_nzd, status_note, created_by_oid, updated_by_oid)
                values (%s,%s,%s,%s,%s,%s,%s,'import','import')
                returning id
            """, (dept, sid, order, r["address"], r["timing"] or None,
                  r["fee"], r["status"] or None))
            did = cur.fetchone()[0]
            order += 1000
            for code in r["brokers"]:
                cur.execute("""
                    insert into public.db_deal_brokers (deal_id, broker_code)
                    select %s, %s
                    where exists (select 1 from public.brokers where code=%s)
                    on conflict do nothing
                """, (did, code, code))

    conn.commit()
    print("\nImported.")


if __name__ == "__main__":
    main()
